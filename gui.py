import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
from tkinter import *
from tkinter import filedialog
from tkinter.filedialog import askdirectory
from sys import exit
from tkinter import messagebox
from tkinter import ttk
import os
from syntax import *
from file_creator import *


      
	








#fonction qui prend en entree une liste de fichiers CLAN et cree les dump et matbase (si la tache n'est pas Hoppy) correspondants
def exec_dump(filelist,sheet,dest,task,fichier):
    j = 0
    for i in range(len(filelist)):
        if (check_for_mistakes(filelist[i])):
            messagebox.showerror("error",filelist[i])
            j = 1
        if (check_subd(filelist[j])):
                j = 1

    if (j == 1):
        messagebox.showerror("Veuillez rentrer à nouveau vos fichiers")
        filelist = []
        return
    else:
        for i in range(len(filelist)):
            addfiletosheet(filelist[i],sheet)
    fichier.save(filename = dest + "/dump.xlsx")
    ref_fichier = openpyxl.load_workbook(dest + "/dump.xlsx")
    ref_sheet = ref_fichier.active
	
    fichier_0 = openpyxl.Workbook()
    sheet_0 = fichier_0.active
    file_list = extract_filenames(ref_sheet)
    sheet_0.append(extract_items(ref_sheet,file_list[0]))
    itemlis = extract_itemsbase(ref_sheet,file_list[0])
    sheet_0.append(second_line(len(itemlis)))
    for i in range(len(file_list)):
        add_file_to_matbase_0(file_list[i],sheet_0,ref_sheet,itemlis)
    for line in sheet_0.iter_rows():
        for cell in line:
            if (cell.value == " 0 0 0 0 0"):
                cell.value = "NR"
    fichier_0.save(dest + "/Matbase_0.xlsx")
    if (task != "Hoppy"): 
        fichier_M = openpyxl.Workbook()
        sheet_M = fichier_M.active
        file_list = extract_filenames(ref_sheet)
        sheet_M.append(extract_items(ref_sheet,file_list[0]))
        itemlis = extract_itemsbase(ref_sheet,file_list[0])
        sheet_M.append(second_line(len(itemlis)))
        for i in range(len(file_list)):
            add_file_to_matbase_M(file_list[i],sheet_M,ref_sheet,itemlis)
        for line in sheet_M.iter_rows():
            for cell in line:
                if (cell.value == " 0 0 0 0 0"):
                    cell.value = "NR"
        fichier_M.save(dest + "/Matbase_M.xlsx")
	
        fichier_D = openpyxl.Workbook()
        sheet_D = fichier_D.active
        file_list = extract_filenames(ref_sheet)
        sheet_D.append(extract_items(ref_sheet,file_list[0]))
        itemlis = extract_itemsbase(ref_sheet,file_list[0])
        sheet_D.append(second_line(len(itemlis)))
        for i in range(len(file_list)):
            add_file_to_matbase_D(file_list[i],sheet_D,ref_sheet,itemlis)
        for line in sheet_0.iter_rows():
            for cell in line:
                if (cell.value == " 0 0 0 0 0"):
                    cell.value = "NR"
        fichier_D.save(dest + "/Matbase_D.xlsx")
        xl_gen(dest+"/dump.xlsx",dest+"/Matbase_0.xlsx",dest,"0")
        xl_gen(dest+"/dump.xlsx",dest+"/Matbase_D.xlsx",dest,"D")
        xl_gen(dest+"/dump.xlsx",dest+"/Matbase_M.xlsx",dest,"M")
        

#fonction permettant a l'utilisateur de selectionner des fichiers a mettre en entree
def browseFiles(filenames,text):
	display = []
	tup_val = filedialog.askopenfilenames(initialdir = "/", title = "Select a File",filetypes = (("CLAN files","*.cha*"),("all files","*.*")))
	for i in range(len(tup_val)):
		filenames.append(tup_val[i])
		display.append(tup_val[i])
	text.insert(END,list(map(removepath,display)))


def clear_list(filelist):
	filelist.clear()
	print(filelist)
	text.delete("1.0",END)
	text.insert("1.0","Fichiers choisis : \n")






def xl_gen(dump_file,mat_file,dest_path,letter):
        d_fichier = openpyxl.load_workbook(dump_file)
        d_sheet = d_fichier.active
        new_file = openpyxl.Workbook()
        new_sheet = new_file.active
        file_list = extract_filenames(d_sheet)
        new_sheet.append(extract_itemsloc(d_sheet,file_list[0]))
        item_lis = extract_itemsbase(d_sheet,file_list[0])
        new_sheet.append(sec_line_loc(len(item_lis)))
        mat_fichier = openpyxl.load_workbook(mat_file)
        m_sheet = mat_fichier.active
        new_file_s = openpyxl.Workbook()
        new_sheet_s = new_file_s.active
        new_sheet_s.append(extract_itemsloc(d_sheet,file_list[0]))
        new_sheet_s.append(sec_line_loc(len(item_lis)))
        for i in range(len(file_list)):
                new_sheet.append(linefiller(file_list[i],item_lis,i+2,m_sheet))
        new_file.save(dest_path + "/Matinter_"+letter+".xlsx")
        for j in range(len(file_list)):
                new_sheet_s.append(linefiller_bis(file_list[j],item_lis,j+2,m_sheet))
        new_file_s.save(dest_path + "/Matinter_bis_"+letter+".xlsx")
        matfinal = openpyxl.Workbook()
        final_sheet = matfinal.active
        final_sheet.append(extract_itemsfinal(d_sheet,file_list[0]))
        final_sheet.append(sec_line_final(len(item_lis)))
        for k in range(len(file_list)):
                final_sheet.append(linefiller_final(file_list[k],item_lis,k+2,m_sheet))
        matfinal.save(dest_path + "/matfinale_"+letter+".xlsx")
        matfinalbis = openpyxl.Workbook()
        bisfinal_sheet = matfinalbis.active
        bisfinal_sheet.append(extract_itemsfinal(d_sheet,file_list[0]))
        bisfinal_sheet.append(sec_line_final(len(item_lis)))
        for n in range(len(file_list)):
                bisfinal_sheet.append(linefiller_finalbis(file_list[n],item_lis,n+2,m_sheet))
        matfinalbis.save(dest_path + "/matfinalebis_"+letter+".xlsx")

