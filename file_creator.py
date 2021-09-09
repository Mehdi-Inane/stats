import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
from syntax import *

#fonction qui enlève le chemin absolu d'un fichier, laissant le chemin relatif.
def removepath(filename):
	lis = re.split("/",filename)
	return lis[-1]



#fonction qui traverse les fichiers CLAN en entrée et retourne uniquement les lignes de code avec les stimulis correspondants.
def detectcodelines(lis):
	newlis = []
	for i in range(len(lis)):
		if (lis[i][0:4] == "%cod"):
			newlis.append(lis[i][5:])
			j = i+1
			while (lis[j][0] == "\t"):
				if (lis[j][2] == "$"):
					newlis.append(lis[j][2:])
				else:
					newlis.append(lis[j])
				j = j+1
		if (lis[i][0:2] == "@G" and (lis[i][4:8] != "XGRE" or lis[i][4:8] != "CMOU")):
			newlis.append(lis[i][4:])
		
	return newlis


#fonction qui divise une ligne de code en plusieurs champs en detectant les caracteres speciaux
def splitspecials(lis):
	newlis = []
	j = 4
	for i in range(len(lis)):
		if (lis[i][0] == "V"):
			if (lis[i][3] == "#"):
				newlis.append(lis[i][0:3])
				if (lis[i][4] == "|"):
					newlis.append("")
				else:
                                        while (lis[i][j]!="|"):
                                                j = j+1
                                        newlis.append(lis[i][4:j])
				newlis.append(lis[i][j+1:])
		
		elif (lis[i][0]=="E"):
			interm = lis[i].split("&")
			newlis.append(interm[0])
			if (lis[i][-1] == "|"):
				newlis.append("")
				if (interm[1] == "|"):
					newlis.append("")
			else:
				interm1 = interm[1].split("|")
				newlis.append(interm1[0])
				newlis.append(interm1[1])
			
		else:
			newlis.append(lis[i])
		j = 4
	return newlis 
		


#fonction qui remplit un dump en prenant en entree le nom du fichier CLAN

def addfiletosheet(name,sheet):
	sheet.append([removepath(name)])
	clanfile = open(name,"r",encoding="utf-8")
	lis = clanfile.readlines()
	code_list = detectcodelines(lis)
	for i in range(len(code_list)):
		formerlist = code_list[i].split(":")
		if (len(formerlist) == 1):
			newstr = removepath(name) + " - " + formerlist[0]
			sheet.append([newstr])
		elif (formerlist[1][0:3] == "IRR"):
			continue
		elif (formerlist[1][-3:] == "/NO"):
			newlis = split_the_nos(formerlist)
			sheet.append(newlis[1:])
		else:	
			newlis = splitspecials(formerlist)
			sheet.append(newlis[1:])
	clanfile.close()


def split_the_nos(lis):
	ret = []
	for i in range(len(lis)):
		if (i == 13 or i == 14):
			ret = ret + [lis[i]]
		else:
			ret = ret + re.split("\||&|#",lis[i])
		
	return ret

#fonction qui remplit la premiere ligne d'un fichier dump               
def extract_items(sheet,filename):
	item_line = ["Nom des stimulis"]
	for cell in sheet["A"]:
		lis = cell.value.split()
		if (len(lis) == 3):
			if (lis[0] == filename):
				item_line = item_line+ 5*[lis[2]]
			else:
				break
		else:
			continue
	return item_line



#fonction qui remplit la premiere ligne d'un matbase
def extract_itemsbase(sheet,filename):
	item_line = ["Nom des stimulis"]
	for cell in sheet["A"]:
		lis = cell.value.split()
		if (len(lis) == 3):
			if (lis[0] == filename):
				item_line = item_line+ [lis[2]]
			else:
				break
		else:
			continue
	return item_line
#fonction qui remplit les deux dernieres colonnes d'un stimuli dans le fichier dump
def extract_c_column(sheet,cell):
	ret_list = []
	upd_list = ""
	nextval = "J" + str(cell.row)
	if (sheet[nextval].value):
                if (sheet[nextval].value[0] != "E"):
                        ret_list.append(sheet[nextval].value)
	nextval = "M" + str(cell.row)
	if (sheet[nextval].value):
                if (sheet[nextval].value[0] != "E"):
                        ret_list.append(sheet[nextval].value)
	nextval = "P" + str(cell.row)
	if (sheet[nextval].value):
                if (sheet[nextval].value[0] != "E"):
                        ret_list.append(sheet[nextval].value)
	nextval = "S" + str(cell.row)
	if (sheet[nextval].value):
                if (sheet[nextval].value[0] != "E"):
                        ret_list.append(sheet[nextval].value)
	nextval = "V" + str(cell.row)
	if (sheet[nextval].value):
                if (sheet[nextval].value[0] != "E"):
                        ret_list.append(sheet[nextval].value)
	for j in range(len(ret_list)):
		if (ret_list[j][0] == "+"):
			upd_list = upd_list + " " + ret_list[j][1:]
		else:
			upd_list = upd_list + " " + ret_list[j]
	if (ret_list == []):
		upd_list = ""
	return upd_list

#fonction qui remplit la seconde ligne d'un fichier dump
def second_line(n):
	ret_list = ["sujet","Type de reponse","[PM]V","[PM]S","[PS]V","[PS]S"]
	pattern = ["C","[PM]V","[PM]S","[PS]V","[PS]S"]
	for i in range(n-2):
		ret_list = ret_list + pattern
	return ret_list


#fonction qui extrait les informations necessaires pour un matbase donne (parametre letter), et un stimuli donne
def extract_stims(sheet,filename,stim,letter):
	boo = 0
	ret_list = []
	ref_cell = filename + " - " + stim + "\n"
	for cell in sheet["A"]:#on cherche la cellule qui débute l'item en question
		split = cell.value.split()
		if (cell.value == ref_cell):
			boo = 1
		if (boo):
			if (cell.value == "CIB" + letter + "/NO" or (cell.value[0]=="C" and cell.value[-1] == letter)):
				ret_list.append(cell.value)
				nextval = "E" + str(cell.row)
				if (sheet[nextval].value):
					ret_list.append(sheet[nextval].value[1:])
				
				else:
					ret_list.append("")
				ret_list.append(extract_c_column(sheet,cell))
				ret_list = ret_list + last_columns(sheet,cell)
				break
			elif (len(split)==3 and split[2] != stim):
				break
			else:
				continue
		
		else:
			continue
	if (ret_list == []):
		ret_list = ["","","","",""]
	return ret_list
		
#creation d'un matbase_0
def subjmatbase_0(sheet,filename,stimlis):
	finalret = []
	finalret.append(filename)
	for i in range(1,len(stimlis)):
		finalret = finalret + extract_stims(sheet,filename,stimlis[i],"0")
	return finalret
#creation d'un matbase_d
def subjmatbase_D(sheet,filename,stimlis):
	finalret = []
	finalret.append(filename)
	for i in range(1,len(stimlis)):
		finalret = finalret + extract_stims(sheet,filename,stimlis[i],"D")
	return finalret
#creation d'un matbase_m
def subjmatbase_M(sheet,filename,stimlis):
	finalret = []
	finalret.append(filename)
	for i in range(1,len(stimlis)):
		finalret = finalret + extract_stims(sheet,filename,stimlis[i],"M")
	return finalret
#remplit les deux dernieres colonnes au nivea d'un fichier dump
def last_columns(sheet,cell):
	i = 1
	nextval = "A" + str(cell.row+i)
	ret_lis = []
	e_word = ""
	c_word = ""
	if (sheet[nextval].value == cell.value):
		while (sheet[nextval].value == cell.value):#on peut check si on a deux trucs
			nextval1 = "X" +str(cell.row)
			nextval2 = "X" + str(cell.row +i)
			if ("PS" in sheet[nextval1].value):
				nextval = "E" + str(cell.row)
				if (sheet[nextval].value):
					e_word = e_word + sheet[nextval].value[1:]
				else:
					e_word = e_word + ""
				c_word = c_word + extract_c_column(sheet,sheet[nextval])
			if ("PS" in sheet[nextval2].value):
				nextval = "E" + str(cell.row+i)
				if (sheet[nextval].value):
					e_word = e_word + sheet[nextval].value[1:]
				else:
					e_word = e_word + ""
				c_word = c_word + extract_c_column(sheet,sheet[nextval])
			i = i+1
			nextval = "A" +str(cell.row + i)
		ret_lis.append(e_word)
		ret_lis.append(c_word)
	else:
		ret_lis.append("")
		ret_lis.append("")
	return ret_lis
#remplit les fichier excel matbase avec la ligne generee par subjmatbase
def add_file_to_matbase_0(filename,sheet,ref_sheet,itemlis):
	sheet.append(subjmatbase_0(ref_sheet,filename,itemlis))

def add_file_to_matbase_D(filename,sheet,ref_sheet,itemlis):
	sheet.append(subjmatbase_D(ref_sheet,filename,itemlis))
def add_file_to_matbase_M(filename,sheet,ref_sheet,itemlis):
	sheet.append(subjmatbase_M(ref_sheet,filename,itemlis))
#prend un fichier dump en entree et extrait le nom des fichiers traites
def extract_filenames(ref_sheet):
	ret_list = []
	for cell in ref_sheet["A"]:
		lis = cell.value.split()
		if (len(lis) == 3):
			if (not(lis[0] in ret_list)):
				ret_list.append(lis[0])
	return ret_list		



#calcul du locus
def extract_itemsloc(sheet,filename):
	item_line = ["Nom des stimulis"]
	for cell in sheet["A"]:
		lis = cell.value.split()
		if (len(lis) == 3):
			if (lis[0] == filename):
				item_line = item_line+ 3*[lis[2]]
			else:
				break
		else:
			continue
	return item_line
def sec_line_loc(n):
	ret_list = ["sujet","Type de reponse","V","S"]
	pattern = ["C","V","S"]
	for i in range(n-2):
		ret_list = ret_list + pattern
	return ret_list



def v_count(line_nb,stim,ref_sheet):
	ret_lis = []
	for cell in ref_sheet[1]:
		if (cell.value == stim):
			#type de réponse
			l = get_column_letter(cell.column)
			ret_lis.append(ref_sheet[l+str(line_nb+1)].value)
			#verbe
			l = get_column_letter(cell.column+1)
			ret_lis.append(letter_count(ref_sheet[l+str(line_nb+1)],"T") + " " + letter_count(ref_sheet[l+str(line_nb+1)],"M"))
			#périphérie
			l = cell.column+2
			ret_lis.append(col_count(l,line_nb+1,ref_sheet))
			return ret_lis

def v_count_bis(line_nb,stim,ref_sheet):
        ret_lis = []
        for cell in ref_sheet[1]:
                if (str(cell.value) == stim):
                #type de réponse
                        l = get_column_letter(cell.column)
                        ret_lis.append(ref_sheet[l+str(line_nb+1)].value)
                #verbe
                        l = get_column_letter(cell.column + 1)
                        ret_lis.append(countbis(ref_sheet[l+str(line_nb+1)]))
                #périphérie
                        l = cell.column +2
                        ret_lis.append(col_countbis(l,line_nb+1,ref_sheet))
                        return ret_lis

def v_count_final(line_nb,stim,ref_sheet):
        ret_lis = []
        for cell in ref_sheet[1]:
                if (str(cell.value) == stim):
                #type de réponse
                        l = get_column_letter(cell.column)
                        ret_lis.append(ref_sheet[l+str(line_nb+1)].value)
                #verbe et périphérie
                        l  = cell.column + 1
                        ret_lis.append(countfinal(l,line_nb+1,ref_sheet))
                        return ret_lis
def v_count_finalbis(line_nb,stim,ref_sheet):
        ret_lis = []
        for cell in ref_sheet[1]:
                if (str(cell.value) == stim):
                        l = get_column_letter(cell.column)
                        ret_lis.append(ref_sheet[l+str(line_nb+1)].value)
                        l = cell.column + 1
                        ret_lis.append(countfinalbis(l,line_nb+1,ref_sheet))
                        return ret_lis



def countfinal(l,line_nb,ref_sheet):
	t = 0
	tcount = 0
	mcount = 0
	tstr = ""
	mstr= ""
	while (t != 4):
		u = get_column_letter(l + t)
		words = (str(ref_sheet[u+str(line_nb)].value)).split()
		for i in range(len(words)):
			if words[i][0] == "T":
				tcount=tcount+1
			if (words[i][0] == "M"):
				mcount = mcount + 1
		t = t +1
	if (tcount !=0):
		tstr = str(tcount) + "T"
		if (mcount != 0):
			tstr = tstr + " "
	if (mcount != 0):
		mstr = str(mcount) + "M"
	if (tcount == 0 and mcount == 0):
		tstr = "NR"
		mstr = ""
	return tstr + mstr
        

def letter_count(cell,letter):
	words = (str(cell.value)).split()
	count = 0
	ret_str = ""
	for i in range(len(words)):
		if words[i][0] == letter:
			count=count+1
	if (count != 0):
		ret_str = str(count) + letter
	return ret_str

def countbis(cell):
        ret_str = ""
        if ("T" in str(cell.value)):
                ret_str = ret_str + "T"
        if ("M" in str(cell.value)):
                ret_str = ret_str + "M"
        return ret_str
def col_countbis(l,line_nb,ref_sheet):
        ret_str = ""
        t = 0
        while (t != 3):
                u = get_column_letter(l+t)
                if ("T" not in ret_str):
                        if ("T" in str(ref_sheet[u + str(line_nb)].value)):
                                ret_str = ret_str + "T"
                if ("M" not in ret_str):
                        if ("M" in str(ref_sheet[u+str(line_nb)].value)):
                                ret_str = ret_str + "M"
                t = t +1
        return ret_str

def countfinalbis(l,line_nb,ref_sheet):
        ret_str= ""
        t = 0
        while (t!=4):
                u = get_column_letter(l+t)
                if ("T" not in ret_str):
                        if ("T" in str(ref_sheet[u+str(line_nb)].value)):
                                ret_str = ret_str + "T"
                if ("M" not in ret_str):
                        if ("M" in str(ref_sheet[u + str(line_nb)].value)):
                                ret_str = ret_str + "M"
                t = t+1
        return ret_str

def col_count(l,line_nb,ref_sheet):
	t = 0
	tcount = 0
	mcount = 0
	tstr = ""
	mstr= ""
	while (t != 3):
		u = get_column_letter(l + t)
		words = (str(ref_sheet[u+str(line_nb)].value)).split()
		for i in range(len(words)):
			if words[i][0] == "T":
				tcount=tcount+1
			if (words[i][0] == "M"):
				mcount = mcount + 1
		t = t +1
	if (tcount !=0):
		tstr = str(tcount) + "T"
		if (mcount != 0):
			tstr = tstr + " "
	if (mcount != 0):
		mstr = str(mcount) + "M"
	if (tcount == 0 and mcount == 0):
		tstr = "NR"
		mstr = ""
	return tstr + mstr
	
def linefiller(filename,stimlist,n,ref_sheet):
	ret_lis = [removepath(filename)]
	for i in range(1,len(stimlist)):
		ret_lis = ret_lis + v_count(n,stimlist[i],ref_sheet)
	return ret_lis	
		
def linefiller_bis(filename,stimlist,n,ref_sheet):
        ret_lis = [removepath(filename)]
        for i in range(1,len(stimlist)):
                ret_lis = ret_lis + v_count_bis(n,stimlist[i],ref_sheet)
        return ret_lis

def linefiller_final(filename,stimlist,n,ref_sheet):
        ret_lis = [removepath(filename)]
        for i in range(1,len(stimlist)):
                ret_lis = ret_lis + v_count_final(n,stimlist[i],ref_sheet)
        return ret_lis
def linefiller_finalbis(filename,stimlist,n,ref_sheet):
        ret_lis = [removepath(filename)]
        for i in range(1,len(stimlist)):
                ret_lis = ret_lis + v_count_finalbis(n,stimlist[i],ref_sheet)
        return ret_lis


def extract_itemsfinal(sheet,filename):
	item_line = ["Nom des stimulis"]
	for cell in sheet["A"]:
		lis = cell.value.split()
		if (len(lis) == 3):
			if (lis[0] == filename):
				item_line = item_line+ 2*[lis[2]]
			else:
				break
		else:
			continue
	return item_line
def sec_line_final(n):
	ret_list = ["sujet","Type de reponse","VSP"]
	pattern = ["C","VSP"]
	for i in range(n-2):
		ret_list = ret_list + pattern
	return ret_list